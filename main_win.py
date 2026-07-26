#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
main_win.py（只做「轉入調劑」版｜先依分頁合併再轉入｜盤點用對帳摘要.xlsx 直接帶入 H 欄數值｜不動格線｜已讀PDF另存/歸檔）

目標：
✅ 只「轉入」(新增明細列)，不動其他表格資料（購買不刪、不改）
✅ 不更新右上彙總（因有公式會自動處理）
✅ 明細只填入（日期/收支原因/支出數量）
✅ H 欄（結存數量）寫公式：上一筆有日期的 H - 本列 F
✅ 不動格線/框線：不 insert_rows、不複製整列格式；只在「下一個空白列」寫入
✅ 先合併再轉入：同一分頁(sheet_id)同一天只處理 1 次
✅ 若同日同原因已存在：不新增列，直接把該列支出(F)加總更新，並重寫該列 H 公式（避免你說的 12/29 少扣/不完全）
✅ 對帳摘要輸出：
   - CSV（保留）
   - XLSX（盤點用：h_final 直接是數字，打開就看到 1462/505.5... noticing）
✅ 已讀PDF另存/歸檔：跑完把 input 的 PDF 搬到輸出資料夾，檔名前加「已讀_」
✅ 避免 MergedCell 寫入錯誤：寫入前檢查

用法：
python main_win.py --input 1141229.pdf --ledger "115鍾new維元社區藥局管制藥品收支結存_D.xlsx" --prefer stream
python main_win.py --input 1141229.xlsx --ledger "115鍾new維元社區藥局管制藥品收支結存_D.xlsx"
（可加 --date 114/12/29 強制指定轉入日期）
"""

import argparse
import re
import shutil
import sys
import warnings
from pathlib import Path
from datetime import datetime, date

# ✅ 新增：防止 BadZipFile / 避開 ~$ 暫存鎖定檔
import os
import zipfile

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

# -----------------------------
# 抑制 ARC4 警告
# -----------------------------
try:
    from cryptography.utils import CryptographyDeprecationWarning
    warnings.filterwarnings("ignore", category=CryptographyDeprecationWarning)
except Exception:
    pass

# -----------------------------
# 可改參數
# -----------------------------
PHARMACY_SUFFIX = "_維元社區藥局"
DEFAULT_REASON = "調劑"
FALLBACK_LEDGER_DATE = "114/12/15"
LEDGER_NAME_KEYWORD = "115鍾new維元社區藥局管制藥品收支結存_D"

SUMMARY_SHEETS_CANDIDATES = ["總表", "Table_1", "SUMMARY", "總帳", "總表單"]

DETAIL_HEADERS = ["日期", "收支原因", "收入數量", "支出數量", "結存數量"]
COL_DATE = "日期"
COL_REASON = "收支原因"
COL_OUT = "支出數量"
COL_BAL = "結存數量"


# ============================================================
# 小工具
# ============================================================

def is_appledouble_sidecar(p: Path) -> bool:
    return p.name.startswith("._")

def looks_like_pdf(p: Path) -> bool:
    try:
        with open(p, "rb") as f:
            return f.read(5) == b"%PDF-"
    except Exception:
        return False

def safe_path_for_write(path: Path) -> Path:
    path = Path(path)
    if not path.exists():
        return path
    stem, suffix = path.stem, path.suffix
    i = 1
    while True:
        cand = path.with_name(f"{stem}_{i}{suffix}")
        if not cand.exists():
            return cand
        i += 1

def _safe_move_file(src: Path, dst_dir: Path) -> Path:
    src = Path(src)
    if not src.exists():
        return src
    dst_dir.mkdir(parents=True, exist_ok=True)
    dst = dst_dir / src.name
    if not dst.exists():
        shutil.move(str(src), str(dst))
        return dst
    stem, suffix = src.stem, src.suffix
    i = 1
    while True:
        dst2 = dst_dir / f"{stem}_{i}{suffix}"
        if not dst2.exists():
            shutil.move(str(src), str(dst2))
            return dst2
        i += 1

def safe_copy_replace(src: Path, dst: Path) -> Path:
    src = Path(src)
    dst = Path(dst)
    if not src.exists():
        raise FileNotFoundError(f"找不到要覆蓋的來源檔：{src}")
    dst.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src, dst)
    return dst

def _norm(x):
    if x is None:
        return ""
    s = str(x)
    for t in [" ", "\n", "\r", "\u3000", "\t"]:
        s = s.replace(t, "")
    return s.strip()

def safe_float(x):
    """轉 float（支援 28.00D：取開頭數字）"""
    if x is None:
        return None
    s = str(x).strip().replace(",", "")
    if s == "":
        return None
    m = re.match(r"^([0-9]+(?:\.[0-9]+)?)", s)
    if not m:
        return None
    try:
        return float(m.group(1))
    except Exception:
        return None

def norm_roc_date(v) -> str:
    """統一成 '114/12/29'"""
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        y = v.year - 1911
        if y <= 0:
            return ""
        return f"{y:03d}/{v.month:02d}/{v.day:02d}"

    s = str(v).strip()
    if not s:
        return ""

    m = re.fullmatch(r"(\d{3})[./](\d{1,2})[./](\d{1,2})", s)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        return f"{y:03d}/{mo:02d}/{d:02d}"

    m = re.fullmatch(r"(\d{3})/(\d{1,2})/(\d{1,2})", s)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        return f"{y:03d}/{mo:02d}/{d:02d}"

    return ""

def parse_report_date(report_path: str) -> str:
    """從抽表 xlsx 文字區找 114.12.29，找不到用 fallback"""
    wb = load_workbook(report_path, data_only=True)
    texts = []
    for ws in wb.worksheets:
        for r in range(1, min(40, ws.max_row) + 1):
            for c in range(1, min(30, ws.max_column) + 1):
                v = ws.cell(r, c).value
                if isinstance(v, str) and v.strip():
                    texts.append(v.strip())
    joined = " ".join(texts)
    m = re.search(r"(\d{3})\.(\d{1,2})\.(\d{1,2})", joined)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        return f"{y:03d}/{mo:02d}/{d:02d}"
    return FALLBACK_LEDGER_DATE

def archive_copy_input(input_path: Path, out_dir: Path):
    """複製輸入檔到 out_dir（保留原檔），檔名前加 已讀_（支援 .pdf / .xlsx）"""
    input_path = Path(input_path)
    if not input_path.exists():
        return
    if input_path.suffix.lower() not in (".pdf", ".xlsx"):
        return

    out_dir.mkdir(parents=True, exist_ok=True)
    new_name = f"已讀_{input_path.stem}{input_path.suffix}"
    dst = out_dir / new_name
    dst = safe_path_for_write(dst)

    shutil.copy2(str(input_path), str(dst))



# ============================================================
# Excel COM：重算公式，讓 openpyxl(data_only=True) 讀得到數值快取
# ============================================================

def excel_recalc_and_save(xlsx_path: Path):
    """用本機 Excel 開啟 → 全部重算 → 存檔關閉"""
    try:
        import win32com.client as win32
    except Exception as e:
        raise RuntimeError("缺少 pywin32：請先 python -m pip install pywin32") from e

    xlsx_path = Path(xlsx_path).resolve()
    excel = win32.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        wb = excel.Workbooks.Open(str(xlsx_path))
        excel.CalculateFullRebuild()
        wb.Save()
        wb.Close(SaveChanges=True)
    finally:
        excel.Quit()


# ============================================================
# ① PDF -> Excel（抽表）
# ============================================================

def extract_tables_to_excel(pdf_path: Path, xlsx_path: Path, pages="all", prefer="auto", strip_text="\n"):
    try:
        import camelot
    except Exception as e:
        raise RuntimeError("尚未安裝 camelot-py，請執行：pip install camelot-py") from e

    if not shutil.which("gs"):
        print("⚠️ 偵測不到 ghostscript(gs)。Windows 建議先用 --prefer stream")

    def read(flavor: str):
        return camelot.read_pdf(str(pdf_path), pages=pages, flavor=flavor, strip_text=strip_text)

    tables_lattice, tables_stream = [], []
    if prefer in ("auto", "lattice"):
        try:
            tables_lattice = read("lattice")
        except Exception as e:
            print(f"⚠️ lattice 讀取失敗：{e}")
    if prefer in ("auto", "stream"):
        try:
            tables_stream = read("stream")
        except Exception as e:
            print(f"⚠️ stream 讀取失敗：{e}")

    if prefer == "auto":
        tables = tables_lattice if len(tables_lattice) >= len(tables_stream) else tables_stream
        chosen = "lattice" if tables is tables_lattice else "stream"
    elif prefer == "lattice":
        tables, chosen = tables_lattice, "lattice"
    else:
        tables, chosen = tables_stream, "stream"

    print(f"📄 PDF: {pdf_path.name}")
    print(f"🧩 抽表策略：{chosen}（抓到 {len(tables)} 個表格）")

    xlsx_path = safe_path_for_write(Path(xlsx_path))

    with pd.ExcelWriter(str(xlsx_path), engine="openpyxl") as writer:
        if len(tables) == 0:
            pd.DataFrame([["未偵測到可抽取的表格。"]]).to_excel(writer, sheet_name="INFO", index=False, header=False)
        else:
            for i, t in enumerate(tables, start=1):
                df = t.df.dropna(how="all")
                df.to_excel(writer, sheet_name=f"Table_{i}"[:31], index=False, header=False)
            try:
                summary = []
                for i, t in enumerate(tables, start=1):
                    summary.append({"sheet": f"Table_{i}", "page": getattr(t, "page", ""), "accuracy": getattr(t, "accuracy", "")})
                pd.DataFrame(summary).to_excel(writer, sheet_name="SUMMARY", index=False)
            except Exception:
                pass

    print(f"✅ 已輸出：{xlsx_path.name}")
    return xlsx_path

def run_pdf_to_excel(input_path: Path, pages="all", prefer="auto") -> Path:
    input_path = Path(input_path).expanduser().resolve()
    if not input_path.exists():
        raise FileNotFoundError(f"找不到輸入檔：{input_path}")

    suf = input_path.suffix.lower()
    if suf == ".xlsx":
        return input_path
    if suf != ".pdf":
        raise ValueError("--input 只支援 .pdf / .xlsx")

    if not looks_like_pdf(input_path):
        raise ValueError(f"這個檔案看起來不是正常 PDF：{input_path.name}")

    xlsx_path = safe_path_for_write(input_path.with_suffix(".xlsx"))
    print("2) PDF tables -> Excel ...")
    return extract_tables_to_excel(input_path, xlsx_path, pages=pages, prefer=prefer)


# ============================================================
# ② 讀報表支出（藥代/使用總量）
# ============================================================

def normalize_code(s: str) -> str:
    # 移除 '-' 等非英數字元，避免像 ERA-4 vs ERA4 對不起來
    return re.sub(r"[^A-Z0-9]", "", _norm(s).upper())

def split_codes(s: str):
    """把 'STI1/STI3' 或含換行斜線拆成多碼"""
    if s is None:
        return []
    t = str(s).upper().replace("\n", "/").replace("\\", "/").replace(" ", "")
    parts = [p.strip() for p in t.split("/") if p.strip()]
    out = []
    for p in parts:
        m = re.findall(r"[A-Z0-9]+", p)
        if m:
            out.append("".join(m))
    seen = set()
    res = []
    for x in out:
        if x not in seen:
            seen.add(x)
            res.append(x)
    return res

def _find_report_header(ws):
    header_row = None
    col_code = None
    col_use = None
    max_r = min(250, ws.max_row)
    max_c = min(200, ws.max_column)

    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            v = _norm(ws.cell(r, c).value)
            if not v:
                continue
            if col_code is None and ("藥代" in v or "代碼" in v):
                col_code = c
                header_row = r
            if col_use is None and ("使用" in v and ("總量" in v or "量" in v)):
                col_use = c
                header_row = r
        if header_row and col_code and col_use:
            return header_row, col_code, col_use

    return None, None, None

def _fallback_qty_scan_row(ws, r, col_code, max_c=200):
    nums = []
    for c in range(1, min(max_c, ws.max_column) + 1):
        if c == col_code:
            continue
        fv = safe_float(ws.cell(r, c).value)
        if fv is not None:
            nums.append(float(fv))
    if not nums:
        return None
    return max(nums)

def read_report_out_qty(report_path: str):
    # ✅ 新增：避開 Excel 暫存鎖定檔 + 檔案不是 zip 的情況，避免 BadZipFile
    base = os.path.basename(report_path)
    if base.startswith("~$"):
        raise ValueError(
            f"❌ 你選到 Excel 暫存鎖定檔：{base}\n"
            f"請關閉 Excel，並改用『沒有 ~$ 開頭』的真正 xlsx 檔。"
        )
    if not zipfile.is_zipfile(report_path):
        raise ValueError(
            f"❌ 檔案不是有效的 Excel xlsx(zip) 格式：\n{report_path}\n"
            f"可能原因：檔案損毀/其實不是 xlsx/仍在被 Excel 佔用尚未完整寫入。"
        )

    wb = load_workbook(report_path, data_only=True)
    table_sheets = [n for n in wb.sheetnames if n.startswith("Table_")]
    sheet_names = table_sheets if table_sheets else [wb.worksheets[0].title]

    out_map = {}
    meta_map = {}

    for sname in sheet_names:
        ws = wb[sname]
        header_row, col_code, col_use = _find_report_header(ws)
        if not (header_row and col_code and col_use):
            continue

        empty_streak = 0
        for r in range(header_row + 1, ws.max_row + 1):
            code_raw = ws.cell(r, col_code).value
            code = normalize_code(code_raw)

            stop = _norm(ws.cell(r, 1).value) + _norm(ws.cell(r, 2).value)
            if "總筆數" in stop:
                break

            if code == "":
                empty_streak += 1
                if empty_streak >= 12:
                    break
                continue
            empty_streak = 0

            qty = safe_float(ws.cell(r, col_use).value)
            if qty is None:
                qty = _fallback_qty_scan_row(ws, r, col_code)
            qty = 0.0 if qty is None else float(qty)

            out_map[code] = out_map.get(code, 0.0) + qty
            if code not in meta_map:
                meta_map[code] = {"sheet": sname, "row": r}

    if not out_map:
        raise ValueError("❌ 報表沒有讀到任何藥代支出資料（請確認 Table_x 有『藥代/使用總量』欄）")
    return out_map, meta_map


# ============================================================
# ③ 總表 -> 對應分頁
# ============================================================

def find_summary_sheet(wb):
    for n in SUMMARY_SHEETS_CANDIDATES:
        if n in wb.sheetnames:
            return wb[n]
    for name in wb.sheetnames:
        if "總" in name:
            return wb[name]
    return wb[wb.sheetnames[0]]

def find_summary_columns(ws):
    header_row = None
    col_id = None
    col_code = None

    for r in range(1, min(80, ws.max_row) + 1):
        for c in range(1, min(150, ws.max_column) + 1):
            v = _norm(ws.cell(r, c).value)
            if v == "編號":
                col_id = c
                header_row = r
            if v == "藥代":
                col_code = c
                header_row = r
        if header_row and col_id and col_code:
            return header_row, col_id, col_code

    raise ValueError(f"❌ 總表({ws.title})找不到表頭『編號/藥代』")

def build_code_to_sheet_map(wb_read, ws_summary):
    header_row, col_id, col_code = find_summary_columns(ws_summary)
    code_to_sheet = {}
    for r in range(header_row + 1, ws_summary.max_row + 1):
        sid = _norm(ws_summary.cell(r, col_id).value)
        codes_cell = ws_summary.cell(r, col_code).value
        if not sid or codes_cell is None:
            continue
        for code in split_codes(str(codes_cell)):
            code_to_sheet[code] = sid
    return code_to_sheet


# ============================================================
# ④ 找明細表 + 寫入（不插入列，不動格線）
# ============================================================

def find_detail_table(ws_read):
    max_r = min(500, ws_read.max_row)
    max_c = min(200, ws_read.max_column)

    for r in range(1, max_r + 1):
        row_vals = [_norm(ws_read.cell(r, c).value) for c in range(1, max_c + 1)]
        if all(h in row_vals for h in DETAIL_HEADERS):
            col_map = {h: row_vals.index(h) + 1 for h in DETAIL_HEADERS}
            col_date = col_map[COL_DATE]
            ok = False
            for rr in range(r + 1, min(r + 12, ws_read.max_row) + 1):
                if norm_roc_date(ws_read.cell(rr, col_date).value):
                    ok = True
                    break
            if ok:
                return r, col_map
    return None, None

def find_prev_data_row(ws_any, header_row, col_date, before_row):
    """找 before_row 上方最近一列『有日期』的列號"""
    for r in range(before_row - 1, header_row, -1):
        if norm_roc_date(ws_any.cell(r, col_date).value):
            return r
    return None

def find_rows_by_date_and_reason(ws_read, header_row, col_map, target_date, reason_text):
    col_date = col_map[COL_DATE]
    col_reason = col_map[COL_REASON]
    tgt = norm_roc_date(target_date) or str(target_date).strip()
    reason_norm = _norm(reason_text)

    rows = []
    max_scan = min(ws_read.max_row, header_row + 8000)
    for r in range(header_row + 1, max_scan + 1):
        dv = norm_roc_date(ws_read.cell(r, col_date).value)
        if not dv or dv != tgt:
            continue
        rv = _norm(ws_read.cell(r, col_reason).value)
        if rv == reason_norm:
            rows.append(r)
    return rows

def set_cell_safe(ws_write, r, c, value=None):
    cell = ws_write.cell(r, c)
    if isinstance(cell, MergedCell):
        return False
    cell.value = value
    return True

def find_next_empty_row(ws_write, ws_read, header_row, col_date):
    """以日期欄為準，找第一個日期欄空白的列"""
    max_scan = min(ws_read.max_row, header_row + 8000)
    for r in range(header_row + 1, max_scan + 1):
        if _norm(ws_read.cell(r, col_date).value) == "" and _norm(ws_write.cell(r, col_date).value) == "":
            return r
    return max_scan + 1

def write_h_formula(ws_write, col_h, col_out, row_here, prev_row):
    """重寫 H 公式：prev_row 的 H - 本列 F"""
    col_h_letter = get_column_letter(col_h)
    col_out_letter = get_column_letter(col_out)
    if prev_row is None:
        formula = f"=-{col_out_letter}{row_here}"
    else:
        formula = f"={col_h_letter}{prev_row}-{col_out_letter}{row_here}"
    set_cell_safe(ws_write, row_here, col_h, formula)
    return f"{col_h_letter}{row_here}"

def append_new_transfer_row(ws_write, ws_read, header_row, col_map, ledger_date, out_qty):
    """新增一列（不插入列、不動格線）"""
    col_date = col_map[COL_DATE]
    col_reason = col_map[COL_REASON]
    col_out = col_map[COL_OUT]
    col_h = col_map[COL_BAL]

    insert_row = find_next_empty_row(ws_write, ws_read, header_row, col_date)
    out_qty = float(out_qty)

    set_cell_safe(ws_write, insert_row, col_date, norm_roc_date(ledger_date) or ledger_date)
    set_cell_safe(ws_write, insert_row, col_reason, DEFAULT_REASON)
    set_cell_safe(ws_write, insert_row, col_out, out_qty)

    prev_row = find_prev_data_row(ws_write, header_row, col_date, insert_row)
    bal_cell = write_h_formula(ws_write, col_h, col_out, insert_row, prev_row)

    return insert_row, bal_cell

def update_existing_dup_row(ws_write, header_row, col_map, dup_row, add_qty):
    """
    ✅ 同日同原因已存在：直接把該列支出(F)加總更新，並重寫該列 H 公式
    """
    col_date = col_map[COL_DATE]
    col_out = col_map[COL_OUT]
    col_h = col_map[COL_BAL]

    add_qty = float(add_qty)

    old_v = safe_float(ws_write.cell(dup_row, col_out).value)
    old_v = 0.0 if old_v is None else float(old_v)
    new_v = old_v + add_qty

    set_cell_safe(ws_write, dup_row, col_out, new_v)

    prev_row = find_prev_data_row(ws_write, header_row, col_date, dup_row)
    bal_cell = write_h_formula(ws_write, col_h, col_out, dup_row, prev_row)

    return dup_row, bal_cell, old_v, new_v


# ============================================================
# ⑤ 對帳摘要輸出：CSV + XLSX(帶入 H 最終數值)
# ============================================================

def write_recon_csv(workdir: Path, ledger_date: str, rows: list):
    ymd = ledger_date.replace("/", "")
    csv_path = workdir / f"對帳摘要_{ymd}.csv"
    pd.DataFrame(rows).to_csv(csv_path, index=False, encoding="utf-8-sig")
    return csv_path

def write_recon_xlsx_values(workdir: Path, ledger_date: str, rows: list, ledger_out_path: Path):
    """
    產生對帳摘要.xlsx（h_final 直接帶入數值）
    前提：ledger_out_path 必須已被 Excel 重算過，才讀得到公式結果的快取數值
    """
    ymd = ledger_date.replace("/", "")
    out_path = workdir / f"對帳摘要_{ymd}.xlsx"

    wb_val = load_workbook(str(ledger_out_path), data_only=True)

    wb = Workbook()
    ws = wb.active
    ws.title = f"對帳摘要_{ymd}"[:31]
    ws.freeze_panes = "A2"

    headers = [
        "sheet_id", "status",
        "out_qty(合計)", "target_row",
        "h_final(數值)", "bal_cell",
        "codes_merged",
        "report_table", "report_row"
    ]
    ws.append(headers)

    for r in rows:
        sheet_id = str(r.get("sheet_id", "")).strip()
        status = r.get("status", "")
        out_qty = r.get("out_qty", "")
        target_row = r.get("target_row", "")
        bal_cell = str(r.get("bal_cell", "")).strip()
        codes_merged = r.get("codes_merged", "")

        h_final = ""
        if status in ("ADDED_NEW", "UPDATED_DUP") and sheet_id and bal_cell and sheet_id in wb_val.sheetnames:
            v = wb_val[sheet_id][bal_cell].value
            fv = safe_float(v)
            h_final = fv if fv is not None else (v if v is not None else "")

        ws.append([
            sheet_id, status,
            out_qty, target_row,
            h_final, bal_cell,
            codes_merged,
            r.get("report_table", ""), r.get("report_row", "")
        ])

    wb.save(out_path)
    return out_path



# ============================================================
# ⑤-2 統計摘要輸出：CSV + XLSX（先統計後轉入用）
# ============================================================

def write_stats_csv(workdir: Path, ledger_date: str, rows: list):
    """
    統計摘要：先把同一分頁(sheet_id)的支出合併好，再輸出一份摘要
    """
    ymd = ledger_date.replace("/", "")
    csv_path = workdir / f"統計摘要_{ymd}.csv"
    pd.DataFrame(rows).to_csv(csv_path, index=False, encoding="utf-8-sig")
    return csv_path

def write_stats_xlsx(workdir: Path, ledger_date: str, rows: list):
    """
    統計摘要.xlsx：純統計（不依賴 Excel 重算）
    """
    ymd = ledger_date.replace("/", "")
    out_path = workdir / f"統計摘要_{ymd}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = f"統計摘要_{ymd}"[:31]
    ws.freeze_panes = "A2"

    headers = [
        "sheet_id",
        "out_qty(合計)",
        "codes_merged",
        "codes_count",
        "report_table",
        "report_row",
    ]
    ws.append(headers)

    for r in rows:
        ws.append([
            str(r.get("sheet_id", "")).strip(),
            r.get("out_qty", ""),
            r.get("codes_merged", ""),
            r.get("codes_count", ""),
            r.get("report_table", ""),
            r.get("report_row", ""),
        ])

    wb.save(out_path)
    return out_path


# ============================================================
# ⑥ 主流程：更新簿冊（先合併再新增/更新）
# ============================================================

def update_ledger_only(report_file: str, ledger_file: str, workdir: Path, forced_date: str = None):
    out_map, meta_map = read_report_out_qty(report_file)
    ledger_date = forced_date or parse_report_date(report_file)

    print(f"📅 轉入日期：{ledger_date}，藥代數：{len(out_map)}")

    wb_read = load_workbook(ledger_file, data_only=True)
    wb_write = load_workbook(ledger_file, data_only=False)

    ws_sum_read = find_summary_sheet(wb_read)
    code_to_sheet = build_code_to_sheet_map(wb_read, ws_sum_read)

    # 先依 sheet_id 合併
    merged = {}  # sheet_id -> {"qty": float, "codes": [], "sources":[("Table_1", 6), ...]}
    missed = []
    recon_rows = []

    for code, qty in out_map.items():
        code_u = normalize_code(code)
        sheet_id = code_to_sheet.get(code_u, "")

        report_table = meta_map.get(code_u, {}).get("sheet", "")
        report_row = meta_map.get(code_u, {}).get("row", "")

        if not sheet_id or sheet_id not in wb_read.sheetnames or sheet_id not in wb_write.sheetnames:
            missed.append(code_u)
            recon_rows.append({
                "sheet_id": sheet_id,
                "status": "MISSING_SHEET",
                "out_qty": float(qty),
                "target_row": "",
                "bal_cell": "",
                "codes_merged": code_u,
                "report_table": report_table,
                "report_row": report_row,
            })
            continue

        if sheet_id not in merged:
            merged[sheet_id] = {"qty": 0.0, "codes": [], "sources": []}

        merged[sheet_id]["qty"] += float(qty)
        merged[sheet_id]["codes"].append(code_u)
        if report_table or report_row:
            merged[sheet_id]["sources"].append((report_table, report_row))

    
    # ✅ 先統計後轉入：先把合併結果輸出成「統計摘要」，再進行寫入
    stats_rows = []
    for sheet_id, info in merged.items():
        report_table, report_row = ("", "")
        if info.get("sources"):
            report_table, report_row = info["sources"][0]
        stats_rows.append({
            "sheet_id": sheet_id,
            "out_qty": float(info.get("qty", 0.0)),
            "codes_merged": "+".join(info.get("codes", [])),
            "codes_count": len(info.get("codes", [])),
            "report_table": report_table,
            "report_row": report_row,
        })

    stats_csv_path = write_stats_csv(workdir, ledger_date, stats_rows)
    stats_xlsx_path = write_stats_xlsx(workdir, ledger_date, stats_rows)


# 逐 sheet_id 寫入（新增或更新既有 dup）
    for sheet_id, info in merged.items():
        qty_sum = float(info["qty"])
        codes_merged = "+".join(info["codes"])

        report_table, report_row = ("", "")
        if info["sources"]:
            report_table, report_row = info["sources"][0]

        ws_read_sheet = wb_read[sheet_id]
        ws_write_sheet = wb_write[sheet_id]

        header_row, col_map = find_detail_table(ws_read_sheet)
        if header_row is None:
            recon_rows.append({
                "sheet_id": sheet_id,
                "status": "NO_DETAIL_HEADER",
                "out_qty": qty_sum,
                "target_row": "",
                "bal_cell": "",
                "codes_merged": codes_merged,
                "report_table": report_table,
                "report_row": report_row,
            })
            continue

        # ✅ 同日同原因已存在：更新最後一筆，不新增列
        dup_rows = find_rows_by_date_and_reason(ws_read_sheet, header_row, col_map, ledger_date, DEFAULT_REASON)
        if dup_rows:
            dup_row = max(dup_rows)
            target_row, bal_cell, old_v, new_v = update_existing_dup_row(
                ws_write_sheet, header_row, col_map, dup_row, qty_sum
            )
            recon_rows.append({
                "sheet_id": sheet_id,
                "status": "UPDATED_DUP",
                "out_qty": qty_sum,
                "target_row": target_row,
                "bal_cell": bal_cell,
                "codes_merged": codes_merged,
                "report_table": report_table,
                "report_row": report_row,
            })
            continue

        # ✅ 否則新增一列（不插入列、不動格線）
        ins_row, bal_cell = append_new_transfer_row(
            ws_write_sheet, ws_read_sheet, header_row, col_map, ledger_date, qty_sum
        )
        recon_rows.append({
            "sheet_id": sheet_id,
            "status": "ADDED_NEW",
            "out_qty": qty_sum,
            "target_row": ins_row,
            "bal_cell": bal_cell,
            "codes_merged": codes_merged,
            "report_table": report_table,
            "report_row": report_row,
        })

    ledger_out = safe_path_for_write(workdir / f"已更新簿冊結存_{ledger_date.replace('/', '')}.xlsx")
    wb_write.save(ledger_out)

    print("🧮 使用 Excel 重算公式中（讓對帳摘要帶入 H 數值）...")
    excel_recalc_and_save(ledger_out)
    print("✅ Excel 重算完成")

    csv_path = write_recon_csv(workdir, ledger_date, recon_rows)
    recon_xlsx_path = write_recon_xlsx_values(workdir, ledger_date, recon_rows, ledger_out)

    if missed:
        print("⚠️ 總表找不到或無法定位分頁：", missed)

    return ledger_date, ledger_out, csv_path, recon_xlsx_path, stats_csv_path, stats_xlsx_path


# ============================================================
# ⑦ 自動挑檔
# ============================================================

def _pick_latest(files):
    files = [Path(f) for f in files if Path(f).is_file() and not is_appledouble_sidecar(Path(f)) and not Path(f).name.startswith("~$")]
    if not files:
        return None
    return max(files, key=lambda p: p.stat().st_mtime)

def auto_select_files(workdir: Path):
    workdir = workdir.resolve()

    pdfs = [p for p in workdir.glob("*.pdf") if p.is_file() and not is_appledouble_sidecar(p) and not p.name.startswith("~$")]
    pdfs += [p for p in workdir.glob("*.PDF") if p.is_file() and not is_appledouble_sidecar(p) and not p.name.startswith("~$")]
    pdfs = [p for p in pdfs if looks_like_pdf(p)]

    xlsxs = [p for p in workdir.glob("*.xlsx") if p.is_file() and not is_appledouble_sidecar(p) and not p.name.startswith("~$")]
    xlsxs += [p for p in workdir.glob("*.XLSX") if p.is_file() and not is_appledouble_sidecar(p) and not p.name.startswith("~$")]

    ledger_candidates = [p for p in xlsxs if "結存" in p.name]
    ledger_named = [p for p in ledger_candidates if LEDGER_NAME_KEYWORD in p.name]
    ledger = _pick_latest(ledger_named) or _pick_latest(ledger_candidates)

    input_file = _pick_latest(pdfs)
    if input_file is None:
        non_ledger_xlsx = [p for p in xlsxs if (ledger is None or p.resolve() != ledger.resolve())]
        input_file = _pick_latest(non_ledger_xlsx)

    return input_file, ledger


def main():
    parser = argparse.ArgumentParser(
        description="只做『調劑轉入』：先依分頁合併；同日同原因已存在就更新支出；H 重寫公式；不插入列不動格線；輸出對帳CSV + 對帳XLSX(帶入 H 數值)；已讀PDF歸檔"
    )
    parser.add_argument("--input", default=None, help="輸入檔：.pdf/.xlsx（不填則自動挑最新）")
    parser.add_argument("--ledger", default=None, help="結存簿 Excel（不填則自動挑最新）")
    parser.add_argument("--pages", default="all", help="Camelot pages，例如 all 或 1,2,3 或 1-end")
    parser.add_argument("--prefer", default="auto", choices=["auto", "lattice", "stream"], help="Camelot flavor")
    parser.add_argument("--date", default=None, help="強制指定轉入日期，例如 114/12/29（不填就從報表抓）")
    args = parser.parse_args()

    workdir = Path.cwd()
    input_path = Path(args.input).expanduser().resolve() if args.input else None
    ledger_path = Path(args.ledger).expanduser().resolve() if args.ledger else None

    if input_path is None or ledger_path is None:
        auto_input, auto_ledger = auto_select_files(workdir)
        if input_path is None:
            input_path = auto_input
        if ledger_path is None:
            ledger_path = auto_ledger

    if not input_path or not Path(input_path).is_file():
        print("❌ 找不到 input 檔案。請把 .PDF 或抽表後 .XLSX 放在此資料夾，或用 --input 指定。")
        sys.exit(2)

    if not ledger_path or not Path(ledger_path).is_file():
        print("❌ 找不到結存簿（ledger）。請把檔名含『結存』的 .XLSX 放在此資料夾，或用 --ledger 指定。")
        sys.exit(2)

    print("📌 使用的 input：", str(input_path))
    print("📌 使用的 ledger：", str(ledger_path))

    report_xlsx = run_pdf_to_excel(Path(input_path), pages=args.pages, prefer=args.prefer)

    ledger_date, ledger_out, csv_path, recon_xlsx_path, stats_csv_path, stats_xlsx_path = update_ledger_only(
        str(report_xlsx), str(ledger_path), workdir, forced_date=args.date
    )

    # 抽表中間檔刪掉
    try:
        p = Path(report_xlsx)
        if p.exists() and p.suffix.lower() == ".xlsx" and p.name.startswith("114"):
            p.unlink()
    except Exception:
        pass

    # 整理輸出到資料夾（✅ 根目錄只留：結存簿 + .py + .bat；✅ 其他全部進「歷史資料匣」）
    out_dir = workdir / f"{ledger_date.replace('/', '')}{PHARMACY_SUFFIX}_報表"
    out_dir.mkdir(parents=True, exist_ok=True)

    # 歷史資料匣（放在 workdir 底下，作為總封存區）
    history_root = workdir / "歷史資料匣"
    history_root.mkdir(parents=True, exist_ok=True)

    # 本次封存資料夾：用 out_dir 的名稱當子資料夾（例如 1150122_維元社區藥局_報表）
    archive_dir = history_root / out_dir.name
    archive_dir.mkdir(parents=True, exist_ok=True)

    # 1) 先把「原始結存簿」備份一份到封存（⚠️ 覆蓋前備份才是真的原始）
    backup_name = f"原始_{Path(ledger_path).name}"
    backup_path = safe_path_for_write(archive_dir / backup_name)
    try:
        shutil.copy2(str(Path(ledger_path)), str(backup_path))
    except Exception as e:
        print("⚠️ 備份原始結存簿失敗：", e)

    # 2) 對帳檔搬入 out_dir（屬於產出）
    _safe_move_file(Path(csv_path), out_dir)
    _safe_move_file(Path(recon_xlsx_path), out_dir)
    _safe_move_file(Path(stats_csv_path), out_dir)
    _safe_move_file(Path(stats_xlsx_path), out_dir)

    # 3) ✅ 已轉入的結存簿要留在外面，並且成為下次帶入的 ledger
    #    做法：把 ledger_out 覆蓋回 ledger_path（外面那份永遠最新）
    try:
        ledger_out_p = Path(ledger_out).resolve()
        ledger_real = Path(ledger_path).resolve()

        if ledger_out_p.exists():
            # 用「安全覆蓋」：把新檔 copy 到原 ledger 位置（保留檔名不變）
            safe_copy_replace(ledger_out_p, ledger_real)

            # 覆蓋成功後，把 ledger_out 這個中間檔刪掉（避免外面多一份）
            try:
                ledger_out_p.unlink()
            except Exception:
                pass

            print("✅ 已將『轉入後結存簿』更新回原檔（外面保留最新）：", str(ledger_real))
        else:
            print("⚠️ 找不到 ledger_out，無法覆蓋回原 ledger：", str(ledger_out_p))
    except Exception as e:
        print("⚠️ 覆蓋回原 ledger 失敗：", e)

    # 4) 輸入檔（pdf/xlsx）封存：做「移動」，根目錄不留
    try:
        in_p = Path(input_path)
        if in_p.exists() and in_p.suffix.lower() in (".pdf", ".xlsx"):
            dst = safe_path_for_write(archive_dir / in_p.name)
            shutil.move(str(in_p), str(dst))
    except Exception as e:
        print("⚠️ 封存輸入檔失敗：", e)

    # 5) 把 out_dir 本身也移入歷史資料匣（根目錄保持乾淨）
    #    若歷史匣內已存在同名資料夾，safe_path_for_write 會自動改名
    try:
        dst_dir = archive_dir  # 預設封存資料夾就是 out_dir 名稱的資料夾
        # out_dir 內容先保留在 out_dir，最後整個搬進 archive_dir 的「內容區」
        # 我們改成：把 out_dir 裡所有檔案/資料夾搬進 archive_dir（避免搬整個資料夾造成自己搬自己）
        for item in list(out_dir.iterdir()):
            try:
                dst_item = safe_path_for_write(dst_dir / item.name)
                shutil.move(str(item), str(dst_item))
            except Exception as e:
                print(f"⚠️ 搬移 out_dir 內容失敗：{item.name} -> {e}")
        # out_dir 變空後刪除
        try:
            out_dir.rmdir()
        except Exception:
            pass
    except Exception as e:
        print("⚠️ 封存 out_dir 失敗：", e)

    # 6) 根目錄只保留：結存簿 + 程式(.py) + 啟動檔(.bat) + 歷史資料匣
    ledger_real = Path(ledger_path).resolve()
    for p in list(workdir.iterdir()):
        if p.name == "歷史資料匣":
            continue
        if p.is_file() and p.suffix.lower() in (".py", ".bat"):
            continue

        # ✅ 115 年結存簿：永遠留在外面（避免被封存進歷史資料匣）
        if p.is_file() and p.name.startswith("115") and ("結存" in p.name):
            continue

        if p.is_file():
            try:
                if p.resolve() == ledger_real:
                    continue
            except Exception:
                if p.name == Path(ledger_path).name:
                    continue
        # 其他一律封存到當次封存資料夾
        try:
            dst = safe_path_for_write(archive_dir / p.name)
            shutil.move(str(p), str(dst))
        except Exception as e:
            print(f"⚠️ 封存失敗：{p.name} -> {e}")

    print(f"📁 歷史資料匣：{history_root}")
    print(f"📦 本次封存資料夾：{archive_dir}")
    print("✅ 根目錄只保留：結存簿 + 程式(.py) + 啟動檔(.bat)")
    print(f"📌 結存簿（被轉入檔，留在原處）：{ledger_path}")
    print(f"📌 原始結存簿備份（封存內）：{backup_path}")

if __name__ == "__main__":


    main()

